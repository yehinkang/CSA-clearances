from datetime import datetime
import json, os, copy
import boto3
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, PatternFill, Font, numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PrintPageSetup
from openpyxl.drawing.image import Image


class GlobalVar:
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))
    
    border_thick = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick'))

    # thick outer, thin inner
    border_top_left = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thin'))

    border_top = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thin'))

    border_top_right = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thin'))

    border_right = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    border_bottom_right = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thick'))

    border_bottom = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick'))

    border_bottom_left = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick'))

    border_left = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))
    
    border_left_closed = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thick'))
    
    border_right_closed = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick'))

    border_top_bottom = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thick'))
    

    border_top_closed = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thin'))
    
    border_bottom_closed = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thick'))

    border_left_right = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    # THICK outer; none inner
    border_thick_top_left = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style='thick'),
        bottom=Side(style=None))

    border_thick_top = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style='thick'),
        bottom=Side(style=None))

    border_thick_top_right = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style=None))

    border_thick_right = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style=None),
        bottom=Side(style=None))

    border_thick_bottom_right = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style=None),
        bottom=Side(style='thick'))

    border_thick_bottom = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thick'))

    border_thick_bottom_left = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thick'))

    border_thick_left = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None))
    
    border_thick_left_closed = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style='thick'),
        bottom=Side(style='thick'))
    
    border_thick_right_closed = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick'))
    
    border_thick_top_bottom = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style='thick'),
        bottom=Side(style='thick'))
    

    border_thick_top_closed = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style=None))
    
    border_thick_bottom_closed = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style=None),
        bottom=Side(style='thick'))

    border_thick_left_right = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style=None),
        bottom=Side(style=None))

    # thin outer; none inner
    border_thin_top_left = Border(
        left=Side(style='thin'),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style=None))

    border_thin_top = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style=None))

    border_thin_top_right = Border(
        left=Side(style=None),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style=None))

    border_thin_right = Border(
        left=Side(style=None),
        right=Side(style='thin'),
        top=Side(style=None),
        bottom=Side(style=None))

    border_thin_bottom_right = Border(
        left=Side(style=None),
        right=Side(style='thin'),
        top=Side(style=None),
        bottom=Side(style='thin'))

    border_thin_bottom = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thin'))

    border_thin_bottom_left = Border(
        left=Side(style='thin'),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thin'))

    border_thin_left = Border(
        left=Side(style='thin'),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None))
    
    border_thin_left_closed = Border(
        left=Side(style='thin'),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style='thin'))
    
    border_thin_right_closed = Border(
        left=Side(style=None),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))
    
    border_thin_top_bottom = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    border_thin_top_closed = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style=None))
    
    border_thin_bottom_closed = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style=None),
        bottom=Side(style='thin'))

    border_thin_left_right = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style=None),
        bottom=Side(style=None))

def create_workbook(workbook_content=None, filename=None, add_footer_index=True, convert_str_to_number=False):
    '''
    INPUTS:
    workbook_content: a list that stores all data to be written into a workbook
                      each list item is a dictionary corresponding to a worksheet
        A worksheet item is composed by data below:
            key='ws_name'; value: name of the worksheet
            key='ws_content'; value: 2D list corresponding to worksheet cells
            key='page_orientation'; value: 0-portrait; 1-landscape
            key='paper_size'; value (Integer): 
                    1 = PAPERSIZE_LETTER
                    2 = PAPERSIZE_LETTER_SMALL
                    3 = PAPERSIZE_TABLOID
                    4 = PAPERSIZE_LEDGER
                    5 = PAPERSIZE_LEGAL
                    6 = PAPERSIZE_STATEMENT
                    7 = PAPERSIZE_EXECUTIVE
                    8 = PAPERSIZE_A3
                    9 = PAPERSIZE_A4
                    10 = PAPERSIZE_A4_SMALL
                    11 = PAPERSIZE_A5

            key='list_img': list of tuples [(anchor_cell, img_path)], e.g. [('A1', path_to_image, width (optional), height (optional))]

            key='cell_range_style': a dictionary that specifies the formatting parameters
                'range_merge': list of tuples; each tuple is composed of:
                    start_row (int), start_column (int),
                        end_row (int), end_column (int), horizontal_align (str, optional)
                    merged cell will be aligned: vertical centered, horizontal per spec

                'range_color':  list of tuples for color range (row_start, col_start, row_end, col_end, rgb_color_code)

                'range_font_bold': list of tuples (row_start, col_start, row_end, col_end)

                'range_font_size': List of tuples for specific font size (row_start, col_start, row_end, col_end, font_size)

                'range_border': list of cell ranges for border (row_min, column_min, row_max, column_max, style_layout (int))
                                style_layout = 0:
                                    Thin borders will be added for all cells in range;
                                    bold borders will be added to the edging sides
                                style_layout = 1: thin border all, outer and inner
                                style_layout = 2: thick outer border only, no inner border
                                style_layout = 3: thin outer border only, no inner border



                'range_align': list of tuples (row_start, col_start, row_end, col_end, alignment_label)
                'range_row_group': list of tuples (row_start, row_end, is_hidden); [is_hidden=False by default]
                'column_width': list of tuples (col_index_based_1, width_in_pt)
                'row_height': list of tuples (row_index_based_1, height_in_pt)
                'range_unwrap': list of tuples (row_start, col_start, row_end, col_end); index is one based
                'range_print_area': list of tuples (row_start, col_start, row_end, col_end)
                

        'footer': list of string, added as notes below table

    filename: name of file with extension .xlsx
    '''

    if workbook_content is None:
        print(f'::: No content is identified for workbook')
        return None

    workbook = Workbook()

    # iterate thru worksheet data
    for ws in workbook_content:
        try:
            ws_name = ws['ws_name']
            worksheet = workbook.create_sheet()
            worksheet.title = ws_name
            ws_cell_value = ws.get('ws_content')

            # set value for each cell
            for i, row in enumerate(ws_cell_value):
                for j, cell_value in enumerate(row):
                    worksheet.cell(row=i + 1, column=j + 1, value=str(cell_value))
                    worksheet.cell(row=i + 1, column=j + 1).alignment = Alignment(wrap_text=True)

                    if convert_str_to_number:
                        if _is_digit(cell_value):
                            if '.' in str(cell_value):
                                worksheet.cell(row=i + 1, column=j + 1).value = float(str(cell_value))
                                worksheet.cell(row=i + 1, column=j + 1).number_format = '0.00'
                            else:
                                worksheet.cell(row=i + 1, column=j + 1).value = int(str(cell_value))
                                worksheet.cell(row=i+1, column=j+1).number_format = numbers.FORMAT_NUMBER

            # formating cells as required: merge/border/color/column-width
            cell_range_style = ws.get('cell_range_style')

            if cell_range_style:
                # unwrap text
                if 'range_unwrap' in cell_range_style:
                    _set_cells_unwrap(worksheet=worksheet, list_cell_range=cell_range_style.get('range_unwrap'))

                # step 1 - align
                if 'range_align' in cell_range_style:
                    _set_format_align(worksheet=worksheet, list_cell_range=cell_range_style['range_align'])

                # Step 2 - color
                if 'range_color' in cell_range_style:
                    _set_format_color(worksheet=worksheet, list_cell_range=cell_range_style['range_color'])

                # Step 3 - font and bold
                if 'range_font_size' in cell_range_style:
                    _set_format_font_size(worksheet=worksheet, list_cell_range=cell_range_style['range_font_size'])

                if 'range_font_bold' in cell_range_style:
                    # will retain font size
                    _set_format_font_bold(worksheet=worksheet, list_cell_range=cell_range_style['range_font_bold'])

                if 'range_merge' in cell_range_style:
                    _set_format_merge(worksheet=worksheet, list_cell_range=cell_range_style['range_merge'])

                if 'range_border' in cell_range_style:
                    _set_format_border(worksheet=worksheet, list_cell_range=cell_range_style['range_border'])

                if 'range_row_group' in cell_range_style:
                    for g in cell_range_style.get('range_row_group'):
                        try:
                            is_hidden = g[2]
                        except:
                            is_hidden = False

                        worksheet.row_dimensions.group(g[0], g[1], hidden=is_hidden)

                if 'column_width' in cell_range_style:
                    for col_width in cell_range_style['column_width']:
                        worksheet.column_dimensions[get_column_letter(col_width[0])].width = col_width[1]
                if 'row_height' in cell_range_style:
                    for row_ht in cell_range_style['row_height']:
                        worksheet.row_dimensions[row_ht[0]].height = row_ht[1]
                
                if 'range_print_area' in cell_range_style:
                    _set_printing_area(worksheet=worksheet, list_cell_range=cell_range_style.get('range_print_area'), 
                                       paper_size=ws.get('paper_size'), orientation=ws.get('page_orientation'))

            # add footer
            list_footer = ws.get('footer')
            row_current = len(ws_cell_value) + 1 # add a blank line between table and footer
            if list_footer:
                for k, footer in enumerate(list_footer):
                    if add_footer_index:
                        footer_row = '    {:d}. {:s}'.format(k+1, footer)
                    else:
                        footer_row = footer

                    worksheet.cell(row=row_current+k+1, column=1, value=str(footer_row))
            
            # insert images
            list_img = ws.get('list_img')
            if list_img:
                _insert_images(worksheet=worksheet, list_img=list_img)

            
        except ValueError as ve:
            print(f'>>> Error: value error detected when creating excel worksheets')
            print(ve.__doc__)
            print(ve.__dict__)

        except Exception as e:
            print(e.__doc__)
            print(f'>>> Error when creating excel worksheets: {e.__doc__}')

    # remove default sheet
    worksheet_default = workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(worksheet_default)
    # rename the first default tab and set as coversheet
    #_add_coversheet(workbook=workbook)
    workbook.save(filename)
    workbook.close()

  
    # return the file name without path info
    return filename


def _set_format_merge(worksheet=None, list_cell_range=None):
    if worksheet is None or list_cell_range is None:
        return

    for cell_range in list_cell_range:
        start_row = cell_range[0]
        start_col = cell_range[1]
        end_row = cell_range[2]
        end_col = cell_range[3]
        worksheet.merge_cells(start_row=start_row, start_column=start_col,
                              end_row=end_row, end_column=end_col)

        try:
            horizontal_align = cell_range[4]
        except:
            horizontal_align = 'left'

        worksheet.cell(row=start_row, column=start_col).alignment = Alignment(wrapText=True, horizontal=horizontal_align, vertical='center')


def _set_format_border(worksheet=None, list_cell_range=None):
    """
    style_layout: fifth item of a cell range
                  0 - thick outer border, thin inner border
                  1 - thin border all, outer and inner
                  2 - thick outer border only, no inner border
                  3 - thin outer border only, no inner border
    """

    if worksheet is None or list_cell_range is None:
        return

    

    for cell_range in list_cell_range:
        row_start = cell_range[0]
        col_start = cell_range[1]
        row_end = cell_range[2]
        col_end = cell_range[3]

        if row_start != row_end and col_start != col_end:
            __set_border_2d_rows_cols(worksheet=worksheet, cell_range_border=cell_range)
        elif row_start == row_end and col_start != col_end:
            __set_border_one_row(worksheet=worksheet, cell_range_border=cell_range)
        elif row_start != row_end and col_start == col_end:
            __set_border_one_col(worksheet=worksheet, cell_range_border=cell_range)
        elif row_start == row_end and col_start == col_end:
            __set_border_one_cell(worksheet=worksheet, cell_range_border=cell_range)
        else:
            return

 
def __set_border_one_cell(worksheet, cell_range_border):
    row = cell_range_border[0]
    col = cell_range_border[1]
    if len(cell_range_border) > 4:
        style_layout = cell_range_border[4]
    else:
        style_layout = 0
    
    cell = worksheet.cell(row=row, column=col)
    if style_layout == 0 or style_layout == 2:
        cell.border = GlobalVar.border_thick

    else:
        cell.border = GlobalVar.border_thin

def __set_border_one_row(worksheet, cell_range_border):
    row = cell_range_border[0]
    col_start = cell_range_border[1]
    col_end = cell_range_border[3]
    if len(cell_range_border) > 4:
        style_layout = cell_range_border[4]
    else:
        style_layout = 0
    
    if style_layout == 0 or style_layout == 1:
        # thin border for all cells
        for col in range(col_start, col_end+1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = GlobalVar.border_thin
        if style_layout == 0:
            # thick outer
            for col in range(col_start, col_end + 1):
                cell = worksheet.cell(row=row, column=col)
                if col == col_start:
                    cell.border = GlobalVar.border_left_closed
                elif col == col_end:
                    cell.border = GlobalVar.border_right_closed
                else:
                    cell.border = GlobalVar.border_top_bottom
    elif style_layout == 2:
        # think outer, empty inner
        for col in range(col_start, col_end + 1):
            cell = worksheet.cell(row=row, column=col)
            if col == col_start:
                cell.border = GlobalVar.border_thick_left_closed
            elif col == col_end:
                cell.border = GlobalVar.border_thick_right_closed
            else:
                cell.border = GlobalVar.border_thick_top_bottom
    elif style_layout == 3:
        # thin outer, empty inner
        for col in range(col_start, col_end + 1):
            cell = worksheet.cell(row=row, column=col)
            if col == col_start:
                cell.border = GlobalVar.border_thin_left_closed
            elif col == col_end:
                cell.border = GlobalVar.border_thin_right_closed
            else:
                cell.border = GlobalVar.border_thin_top_bottom


def __set_border_one_col(worksheet, cell_range_border):
    row_start = cell_range_border[0]
    col = cell_range_border[1]
    row_end = cell_range_border[2]
    
    if len(cell_range_border) > 4:
        style_layout = cell_range_border[4]
    else:
        style_layout = 0

    if style_layout == 0 or style_layout == 1:
        # thin border for all cells
        for row in range(row_start, row_end+1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = GlobalVar.border_thin
        if style_layout == 0:
            # thick outer
            for row in range(row_start, row_end + 1):
                cell = worksheet.cell(row=row, column=col)
                if row == row_start:
                    cell.border = GlobalVar.border_top_closed
                elif row == row_end:
                    cell.border = GlobalVar.border_bottom_closed
                else:
                    cell.border = GlobalVar.border_left_right
    elif style_layout == 2:
        # think outer, empty inner
        for row in range(row_start, row_end + 1):
            cell = worksheet.cell(row=row, column=col)
            if row == row_start:
                cell.border = GlobalVar.border_thick_top_closed
            elif row == row_end:
                cell.border = GlobalVar.border_thick_bottom_closed
            else:
                cell.border = GlobalVar.border_thick_left_right
    elif style_layout == 3:
        # thin outer, empty inner
        for row in range(row_start, row_end + 1):
            cell = worksheet.cell(row=row, column=col)
            if row == row_start:
                cell.border = GlobalVar.border_thin_top_closed
            elif row == row_end:
                cell.border = GlobalVar.border_thin_bottom_closed
            else:
                cell.border = GlobalVar.border_thin_left_right
    

def __set_border_2d_rows_cols(worksheet, cell_range_border):

    row_start = cell_range_border[0]
    col_start = cell_range_border[1]
    row_end = cell_range_border[2]
    col_end = cell_range_border[3]

    if len(cell_range_border) >4:
        # style_layout specified
        style_layout = cell_range_border[4]
    else:
        style_layout = 0

    
    if style_layout == 0 or style_layout == 1:
        # think border for all cells
        for row in range(row_start, row_end+1):
            for col in range(col_start, col_end+1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = GlobalVar.border_thin

        if style_layout == 0:
            __set_border_2d_rows_cols_outer(worksheet=worksheet, row_start=row_start, row_end=row_end, 
                                            col_start=col_start, col_end=col_end, style_layout=style_layout)
    elif style_layout == 2 or style_layout == 3:
        __set_border_2d_rows_cols_outer(worksheet=worksheet, row_start=row_start, row_end=row_end, 
                                            col_start=col_start, col_end=col_end, style_layout=style_layout)

def __set_border_2d_rows_cols_outer(worksheet, row_start, row_end, col_start, col_end, style_layout):
    """
    Set outer border
    """
    if style_layout == 0:
        # thick outer, thin inner
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                cell = worksheet.cell(row=i, column=j)
                if i == row_start and j == col_start:
                    # upper left corner
                    cell.border = GlobalVar.border_top_left
                elif i == row_start and j < col_end:
                    cell.border = GlobalVar.border_top
                elif i == row_start and j == col_end:
                    cell.border = GlobalVar.border_top_right
                elif i == row_end and j == col_start:
                    cell.border = GlobalVar.border_bottom_left
                elif i == row_end and j < col_end:
                    cell.border = GlobalVar.border_bottom
                elif i == row_end and j == col_end:
                    cell.border = GlobalVar.border_bottom_right
                elif j == col_start and i < row_end:
                    cell.border = GlobalVar.border_left
                elif j == col_end and i < row_end:
                    cell.border = GlobalVar.border_right
    
    elif style_layout == 2:
        # thick outer, empty inner
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                cell = worksheet.cell(row=i, column=j)
                if i == row_start and j == col_start:
                    # upper left corner
                    cell.border = GlobalVar.border_thick_top_left
                elif i == row_start and j < col_end:
                    cell.border = GlobalVar.border_thick_top
                elif i == row_start and j == col_end:
                    cell.border = GlobalVar.border_thick_top_right
                elif i == row_end and j == col_start:
                    cell.border = GlobalVar.border_thick_bottom_left
                elif i == row_end and j < col_end:
                    cell.border = GlobalVar.border_thick_bottom
                elif i == row_end and j == col_end:
                    cell.border = GlobalVar.border_thick_bottom_right
                elif j == col_start and i < row_end:
                    cell.border = GlobalVar.border_thick_left
                elif j == col_end and i < row_end:
                    cell.border = GlobalVar.border_thick_right
    elif style_layout == 3:
        # thin outer, empty inner
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                cell = worksheet.cell(row=i, column=j)
                if i == row_start and j == col_start:
                    # upper left corner
                    cell.border = GlobalVar.border_thin_top_left
                elif i == row_start and j < col_end:
                    cell.border = GlobalVar.border_thin_top
                elif i == row_start and j == col_end:
                    cell.border = GlobalVar.border_thin_top_right
                elif i == row_end and j == col_start:
                    cell.border = GlobalVar.border_thin_bottom_left
                elif i == row_end and j < col_end:
                    cell.border = GlobalVar.border_thin_bottom
                elif i == row_end and j == col_end:
                    cell.border = GlobalVar.border_thin_bottom_right
                elif j == col_start and i < row_end:
                    cell.border = GlobalVar.border_thin_left
                elif j == col_end and i < row_end:
                    cell.border = GlobalVar.border_thin_right

def _set_format_color(worksheet=None, list_cell_range=None):

    if worksheet is None or list_cell_range is None:
        return

    #print(f'list_color_range={list_cell_range}')
    for cell_range in list_cell_range:
        color_rgb = cell_range[4]
        pattern_fill = PatternFill('solid', fgColor=color_rgb)

        for i in range(cell_range[0], cell_range[2]+1):
            for j in range(cell_range[1], cell_range[3]+1):
                cell = worksheet.cell(row=i, column=j)
                cell.fill = pattern_fill

        #rows = worksheet.iter_rows(min_row=cell_range[0], max_row=cell_range[2], max_col=cell_range[3])
        #for row in rows:
        #    for cell in row:
        #        cell.fill = pattern_fill


def _set_cells_unwrap(worksheet=None, list_cell_range=None):
    if worksheet is None or list_cell_range is None:
        return

    for cell_range in list_cell_range:
        row_start = cell_range[0]
        col_start = cell_range[1]
        row_end = cell_range[2]
        col_end = cell_range[3]
        for idx_row in range(row_end - row_start + 1): # including start and end
            for idx_col in range(col_end - col_start + 1):
                worksheet.cell(row=row_start + idx_row, column=col_start + idx_col). \
                    alignment = Alignment(wrapText=False)


def _set_format_align(worksheet=None, list_cell_range=None):
    if worksheet is None or list_cell_range is None:
        return
    for cell_range in list_cell_range:
        # cell_ranger: row_start, col_start, row_end, col_end, alignment_label
        row_start = cell_range[0]
        col_start = cell_range[1]
        row_end = cell_range[2]
        col_end = cell_range[3]
        align = cell_range[4]
        for idx_row in range(row_end - row_start + 1): # including start and end
            for idx_col in range(col_end - col_start + 1):
                worksheet.cell(row=row_start + idx_row, column=col_start + idx_col). \
                    alignment = Alignment(horizontal=align, vertical='center')


def _set_format_font_size(worksheet=None, list_cell_range=None):
    if worksheet is None or list_cell_range is None:
        return

    for cell_range in list_cell_range:
        font_style = Font(size=cell_range[4])
        for i in range(cell_range[0], cell_range[2]+1):
            for j in range(cell_range[1], cell_range[3]+1):
                worksheet.cell(row=i, column=j).font = font_style


def _set_format_font_bold(worksheet=None, list_cell_range=None):
    if worksheet is None or list_cell_range is None:
        return

    for cell_range in list_cell_range:
        # cell_ranger: row_start, col_start, row_end, col_end
        row_start = cell_range[0]
        col_start = cell_range[1]
        row_end = cell_range[2]
        col_end = cell_range[3]
        for idx_row in range(row_end - row_start + 1): # including start and end
            for idx_col in range(col_end - col_start + 1):
                cell = worksheet.cell(row=row_start + idx_row, column=col_start + idx_col)
                # retain font size
                font_size_ori = cell.font.size
                font_style = Font(size=font_size_ori, bold=True)
                worksheet.cell(row=row_start + idx_row, column=col_start + idx_col).font = font_style


def _set_printing_area(worksheet=None, list_cell_range=None, orientation=None, paper_size=None):
    list_range_print = []
    for cell_range in list_cell_range:
        first_row = cell_range[0]
        first_col = cell_range[1]
        last_row = cell_range[2]
        last_col = cell_range[3]
        print_area = f'{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}'
        list_range_print.append(print_area)
    
    # add the printing area specs
    worksheet.print_area = list_range_print

    # scale fit to one page width
    #worksheet.page_setup.fitToWidth = True
    
    p = PrintPageSetup(worksheet=worksheet, fitToWidth=1, scale=30)
    if paper_size:
        p.paperSize = paper_size # 3 = Tabloid
    
    if orientation:
        p.orientation = orientation

    p.autoPageBreaks = False
    p.fitToWidth = 1
    p.scale = None
    
    worksheet.page_setup = p
    print('Completed setting priting area.')


def _insert_images(worksheet, list_img):
    for img in list_img:
        obj_img = Image(img[0])
        obj_img.anchor = img[1]
        if len(img) > 2:
            obj_img.width = img[2]
        if len(img) > 3:
            obj_img.height = img[3]
        obj_img.top = 5
        obj_img.left = 5
        worksheet.add_image(obj_img)


def _is_digit(data):
    if type(data) is not str:
        data_new = copy.copy(str(data))
    else:
        data_new = copy.copy(data)

    if data_new.startswith('-'):
        data_new = data_new[1:]

    is_digit = data_new.replace('.', '', 1).isdigit()

    return is_digit